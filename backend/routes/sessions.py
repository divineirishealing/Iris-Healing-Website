from fastapi import APIRouter, HTTPException, UploadFile, File
from models import Session, SessionCreate
from typing import List, Optional
import io
import os
import uuid
from motor.motor_asyncio import AsyncIOMotorClient
from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent.parent
load_dotenv(ROOT_DIR / '.env')

router = APIRouter(prefix="/api/sessions", tags=["Sessions"])

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

@router.get("", response_model=List[Session])
async def get_sessions(visible_only: Optional[bool] = None):
    query = {}
    if visible_only:
        query["visible"] = True
    sessions = await db.sessions.find(query).sort("order", 1).to_list(100)
    return [Session(**session) for session in sessions]

@router.post("/upload-excel")
async def upload_sessions_excel(file: UploadFile = File(...)):
    if not file.filename.endswith(('.xlsx', '.xls', '.xlss')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are supported")

    from openpyxl import load_workbook
    content = await file.read()
    wb = load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Excel file has no data rows")

    raw_headers = [str(h).strip().lower() if h else '' for h in rows[0]]

    # Flexible column matching — matches partial keywords
    col_indices = {}
    for idx, h in enumerate(raw_headers):
        if not h:
            continue
        if 'title' in h or 'session name' in h or 'name' in h:
            if 'title' not in col_indices:  # first match wins
                col_indices['title'] = idx
        elif 'description' in h or 'desc' in h:
            col_indices['description'] = idx
        elif 'type' in h or 'mode' in h:
            col_indices['session_type'] = idx
        elif 'duration' in h or 'time' in h or 'mins' in h or 'minutes' in h:
            col_indices['duration'] = idx
        elif 'inr' in h:
            col_indices['price_inr'] = idx
        elif 'usd' in h:
            col_indices['price_usd'] = idx
        elif 'aed' in h:
            col_indices['price_aed'] = idx

    if 'title' not in col_indices:
        raise HTTPException(status_code=400, detail=f"Could not find a title column. Found headers: {[str(h).strip() for h in rows[0] if h]}")

    created, updated, skipped = 0, 0, 0
    count = await db.sessions.count_documents({})

    for row in rows[1:]:
        title_val = row[col_indices['title']] if col_indices.get('title') is not None and col_indices['title'] < len(row) else None
        if not title_val or not str(title_val).strip():
            skipped += 1
            continue

        title = str(title_val).strip()

        def get_val(field, _row=row):
            idx = col_indices.get(field)
            if idx is None or idx >= len(_row):
                return None
            return _row[idx]

        raw_type = str(get_val('session_type') or '').strip().lower()
        if 'online' in raw_type and 'offline' in raw_type:
            session_mode = 'both'
        elif 'offline' in raw_type or 'in-person' in raw_type or 'in person' in raw_type:
            session_mode = 'offline'
        elif 'both' in raw_type:
            session_mode = 'both'
        else:
            session_mode = 'online'

        dur_val = get_val('duration')
        duration = ''
        if dur_val and str(dur_val).strip():
            try:
                duration = f"{int(float(str(dur_val).strip()))} minutes"
            except (ValueError, TypeError):
                duration = str(dur_val).strip()

        def parse_price(field):
            v = get_val(field)
            if v is None or str(v).strip() == '':
                return 0.0
            try:
                return float(str(v).replace(',', '').strip())
            except (ValueError, TypeError):
                return 0.0

        update_fields = {'title': title}
        desc = get_val('description')
        if desc and str(desc).strip():
            update_fields['description'] = str(desc).strip()
        if session_mode:
            update_fields['session_mode'] = session_mode
        if duration:
            update_fields['duration'] = duration

        p_inr = parse_price('price_inr')
        p_usd = parse_price('price_usd')
        p_aed = parse_price('price_aed')
        if p_inr:
            update_fields['price_inr'] = p_inr
        if p_usd:
            update_fields['price_usd'] = p_usd
        if p_aed:
            update_fields['price_aed'] = p_aed

        existing = await db.sessions.find_one({"title": {"$regex": f"^{title}$", "$options": "i"}})
        if existing:
            await db.sessions.update_one({"id": existing["id"]}, {"$set": update_fields})
            updated += 1
        else:
            session_data = {
                'id': str(uuid.uuid4()),
                'title': title,
                'description': update_fields.get('description', ''),
                'image': '',
                'price_usd': p_usd,
                'price_inr': p_inr,
                'price_eur': 0.0,
                'price_gbp': 0.0,
                'price_aed': p_aed,
                'duration': duration or '60 minutes',
                'session_mode': session_mode,
                'available_dates': [],
                'time_slots': [],
                'testimonial_text': '',
                'visible': True,
                'order': count + created,
            }
            await db.sessions.insert_one(session_data)
            created += 1

    wb.close()
    return {"message": f"Upload complete: {created} created, {updated} updated, {skipped} skipped", "created": created, "updated": updated, "skipped": skipped}

@router.get("/{session_id}", response_model=Session)
async def get_session(session_id: str):
    session = await db.sessions.find_one({"id": session_id})
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    return Session(**session)

@router.post("", response_model=Session)
async def create_session(session: SessionCreate):
    count = await db.sessions.count_documents({})
    data = session.dict()
    data.pop("order", None)
    session_obj = Session(**data, order=count)
    await db.sessions.insert_one(session_obj.dict())
    return session_obj

@router.put("/{session_id}", response_model=Session)
async def update_session(session_id: str, session: SessionCreate):
    existing = await db.sessions.find_one({"id": session_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Session not found")
    update_data = {k: v for k, v in session.dict().items() if v is not None}
    await db.sessions.update_one({"id": session_id}, {"$set": update_data})
    updated = await db.sessions.find_one({"id": session_id})
    return Session(**updated)

@router.patch("/{session_id}/visibility")
async def toggle_visibility(session_id: str, data: dict):
    existing = await db.sessions.find_one({"id": session_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Session not found")
    await db.sessions.update_one({"id": session_id}, {"$set": {"visible": data.get("visible", True)}})
    return {"message": "Visibility updated", "visible": data.get("visible", True)}

@router.patch("/reorder")
async def reorder_sessions(data: dict):
    order_list = data.get("order", [])
    for idx, session_id in enumerate(order_list):
        await db.sessions.update_one({"id": session_id}, {"$set": {"order": idx}})
    return {"message": "Sessions reordered successfully"}

@router.delete("/{session_id}")
async def delete_session(session_id: str):
    result = await db.sessions.delete_one({"id": session_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Session not found")
    return {"message": "Session deleted successfully"}
