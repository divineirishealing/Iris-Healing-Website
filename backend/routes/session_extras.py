from fastapi import APIRouter, HTTPException, UploadFile, File
from models import SessionTestimonial, SessionTestimonialCreate, SessionQuestion, SessionQuestionCreate
from typing import List, Optional
import os
import io
import uuid
from datetime import datetime, timezone
from motor.motor_asyncio import AsyncIOMotorClient
from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent.parent
load_dotenv(ROOT_DIR / '.env')

router = APIRouter(prefix="/api/session-extras", tags=["SessionExtras"])

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# ========== TESTIMONIALS ==========

@router.get("/testimonials", response_model=List[SessionTestimonial])
async def get_session_testimonials(session_id: Optional[str] = None):
    query = {}
    if session_id:
        query["session_id"] = session_id
    items = await db.session_testimonials.find(query).sort("order", 1).to_list(200)
    return [SessionTestimonial(**t) for t in items]

@router.post("/testimonials", response_model=SessionTestimonial)
async def create_session_testimonial(data: SessionTestimonialCreate):
    count = await db.session_testimonials.count_documents({"session_id": data.session_id})
    data_dict = data.dict()
    data_dict["order"] = count  # Override order with actual count
    obj = SessionTestimonial(**data_dict)
    await db.session_testimonials.insert_one(obj.dict())
    return obj

@router.put("/testimonials/{tid}", response_model=SessionTestimonial)
async def update_session_testimonial(tid: str, data: SessionTestimonialCreate):
    existing = await db.session_testimonials.find_one({"id": tid})
    if not existing:
        raise HTTPException(status_code=404, detail="Testimonial not found")
    update = {k: v for k, v in data.dict().items() if v is not None}
    await db.session_testimonials.update_one({"id": tid}, {"$set": update})
    updated = await db.session_testimonials.find_one({"id": tid})
    return SessionTestimonial(**updated)

@router.delete("/testimonials/{tid}")
async def delete_session_testimonial(tid: str):
    result = await db.session_testimonials.delete_one({"id": tid})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Testimonial not found")
    return {"message": "Deleted"}

@router.post("/testimonials/upload-excel")
async def upload_session_testimonials(file: UploadFile = File(...)):
    from openpyxl import load_workbook
    content = await file.read()
    wb = load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="No data rows")
    headers = [str(h).strip().lower() if h else '' for h in rows[0]]
    col = {}
    for idx, h in enumerate(headers):
        if 'session' in h and ('id' in h or 'title' in h or 'name' in h):
            col['session'] = idx
        elif 'client' in h or 'name' in h:
            if 'session' not in h:
                col['name'] = idx
        elif 'testimonial' in h or 'text' in h or 'review' in h or 'feedback' in h:
            col['text'] = idx
    created = 0
    for row in rows[1:]:
        def val(key):
            i = col.get(key)
            if i is None or i >= len(row):
                return ''
            return str(row[i]).strip() if row[i] else ''
        text = val('text')
        if not text:
            continue
        session_ref = val('session')
        if session_ref:
            session = await db.sessions.find_one({"$or": [{"id": session_ref}, {"title": {"$regex": f"^{session_ref}$", "$options": "i"}}]})
            session_id = session["id"] if session else ""
        else:
            session_id = ""
        count = await db.session_testimonials.count_documents({})
        obj = SessionTestimonial(session_id=session_id, client_name=val('name'), text=text, order=count)
        await db.session_testimonials.insert_one(obj.dict())
        created += 1
    wb.close()
    return {"message": f"Uploaded {created} testimonials", "created": created}

# ========== QUESTIONS ==========

@router.get("/questions", response_model=List[SessionQuestion])
async def get_questions(session_id: Optional[str] = None, unreplied_only: Optional[bool] = None):
    query = {}
    if session_id:
        query["session_id"] = session_id
    if unreplied_only:
        query["replied"] = False
    items = await db.session_questions.find(query).sort("created_at", -1).to_list(200)
    return [SessionQuestion(**q) for q in items]

@router.post("/questions", response_model=SessionQuestion)
async def submit_question(data: SessionQuestionCreate):
    obj = SessionQuestion(**data.dict())
    await db.session_questions.insert_one(obj.dict())
    return obj

@router.put("/questions/{qid}/reply")
async def reply_to_question(qid: str, data: dict):
    existing = await db.session_questions.find_one({"id": qid})
    if not existing:
        raise HTTPException(status_code=404, detail="Question not found")
    reply_text = data.get("reply", "")
    send_email = data.get("send_email", False)
    await db.session_questions.update_one({"id": qid}, {"$set": {
        "reply": reply_text,
        "replied": True,
        "replied_at": datetime.now(timezone.utc).isoformat()
    }})
    if send_email and existing.get("email"):
        try:
            from routes.emails import send_email
            await send_email(
                to=existing["email"],
                subject=f"Re: Your Question about {existing.get('session_id', 'our sessions')}",
                html=f"<p>Dear {existing.get('name', 'Seeker')},</p><p>Thank you for your question.</p><p><strong>Your question:</strong> {existing['question']}</p><p><strong>Our response:</strong> {reply_text}</p><p>With love and light,<br>Divine Iris Healing</p>"
            )
        except Exception as e:
            print(f"Email send failed: {e}")
    return {"message": "Reply saved", "email_sent": send_email}

@router.delete("/questions/{qid}")
async def delete_question(qid: str):
    result = await db.session_questions.delete_one({"id": qid})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Question not found")
    return {"message": "Deleted"}

# ========== UNIFIED CALENDAR ==========

@router.get("/calendar")
async def get_calendar():
    cal = await db.session_calendar.find_one({"id": "global_calendar"})
    if not cal:
        default = {
            "id": "global_calendar",
            "available_dates": [],
            "time_slots": ["10:00 AM", "2:00 PM", "5:00 PM"],
            "blocked_until": "",
            "min_advance_days": 7,
            "max_future_months": 3,
            "block_weekends": True,
            "rules_note": "",
        }
        await db.session_calendar.insert_one(default)
        return {k: v for k, v in default.items() if k != '_id'}
    return {k: v for k, v in cal.items() if k != '_id'}

@router.put("/calendar")
async def update_calendar(data: dict):
    data.pop("_id", None)
    data.pop("id", None)
    await db.session_calendar.update_one(
        {"id": "global_calendar"},
        {"$set": data},
        upsert=True
    )
    updated = await db.session_calendar.find_one({"id": "global_calendar"})
    return {k: v for k, v in updated.items() if k != '_id'}
